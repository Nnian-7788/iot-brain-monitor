#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel随机数生成器
功能：在Excel文件的指定列生成指定行数的随机数，支持自定义随机数范围

使用方法：
    python excel_random_generator.py

依赖：
    pip install openpyxl
"""

import os
import random
from openpyxl import load_workbook
from openpyxl.styles import numbers


def get_user_input(prompt: str, input_type: type = str, default=None):
    """
    获取用户输入，支持类型转换和默认值
    
    Args:
        prompt: 提示信息
        input_type: 输入类型 (str, int, float)
        default: 默认值
    
    Returns:
        转换后的用户输入值
    """
    while True:
        try:
            user_input = input(prompt).strip()
            if not user_input and default is not None:
                return default
            if input_type == str:
                return user_input
            elif input_type == int:
                return int(user_input)
            elif input_type == float:
                return float(user_input)
            else:
                return user_input
        except ValueError:
            print(f"❌ 输入格式错误，请输入有效的{input_type.__name__}类型值")


def validate_file_path(file_path: str) -> bool:
    """
    验证文件路径是否存在且为Excel文件
    
    Args:
        file_path: 文件路径
    
    Returns:
        是否有效
    """
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        return False
    
    if not file_path.lower().endswith(('.xlsx', '.xlsm')):
        print("❌ 文件格式错误，请提供 .xlsx 或 .xlsm 格式的Excel文件")
        return False
    
    return True


def generate_random_numbers(min_val: float, max_val: float, count: int, is_integer: bool = False):
    """
    生成随机数列表
    
    Args:
        min_val: 最小值
        max_val: 最大值
        count: 数量
        is_integer: 是否为整数
    
    Returns:
        随机数列表
    """
    if is_integer:
        return [random.randint(int(min_val), int(max_val)) for _ in range(count)]
    else:
        return [round(random.uniform(min_val, max_val), 2) for _ in range(count)]


def column_letter_to_index(column_input: str) -> int:
    """
    将列字母或数字转换为列索引（从1开始）
    
    Args:
        column_input: 列字母（如A, B, AA）或数字
    
    Returns:
        列索引（从1开始）
    """
    column_input = column_input.strip().upper()
    
    # 如果是纯数字，直接返回
    if column_input.isdigit():
        return int(column_input)
    
    # 如果是字母，转换为数字
    result = 0
    for char in column_input:
        if 'A' <= char <= 'Z':
            result = result * 26 + (ord(char) - ord('A') + 1)
        else:
            raise ValueError(f"无效的列标识: {column_input}")
    
    return result


def index_to_column_letter(index: int) -> str:
    """
    将列索引转换为列字母
    
    Args:
        index: 列索引（从1开始）
    
    Returns:
        列字母
    """
    result = ""
    while index > 0:
        index -= 1
        result = chr(ord('A') + index % 26) + result
        index //= 26
    return result


def find_append_start_row(sheet, columns):
    """
    根据指定列中已有数据，计算自动追加时的起始行
    """
    max_last_row = 0
    sheet_max_row = sheet.max_row or 1
    for col in columns:
        for row in range(sheet_max_row, 0, -1):
            if sheet.cell(row=row, column=col).value not in (None, ""):
                max_last_row = max(max_last_row, row)
                break
    return max_last_row + 1 if max_last_row > 0 else 1


def process_excel_file(file_path: str, columns, start_row: int, row_count: int,
                       min_val: float, max_val: float, is_integer: bool,
                       output_path: str = None, auto_append: bool = False):
    """
    处理Excel文件，在指定列生成随机数
    
    Args:
        file_path: Excel文件路径
        column: 列索引（从1开始）
        start_row: 起始行（从1开始）
        row_count: 生成行数
        min_val: 随机数最小值
        max_val: 随机数最大值
        is_integer: 是否为整数
        output_path: 输出文件路径（默认覆盖原文件）
    """
    try:
        # 加载工作簿
        print(f"\n📂 正在加载文件: {file_path}")
        workbook = load_workbook(file_path)

        # 获取活动工作表
        sheet = workbook.active
        print(f"📊 工作表名称: {sheet.title}")

        # 自动计算起始行（接在已有数据后）
        if auto_append:
            start_row = find_append_start_row(sheet, columns)
            print(f"📍 已选择自动追加模式，将从第 {start_row} 行开始写入")

        col_letters = [index_to_column_letter(c) for c in columns]
        print(f"\n✍️ 正在写入数据到列: {', '.join(col_letters)}")

        for col in columns:
            random_numbers = generate_random_numbers(min_val, max_val, row_count, is_integer)
            col_letter = index_to_column_letter(col)
            for i, num in enumerate(random_numbers):
                current_row = start_row + i
                cell = sheet.cell(row=current_row, column=col)
                cell.value = num
                if is_integer:
                    cell.number_format = numbers.FORMAT_NUMBER
                else:
                    cell.number_format = numbers.FORMAT_NUMBER_00

        # 确定输出路径
        if output_path is None:
            output_path = file_path

        # 保存文件
        workbook.save(output_path)
        print(f"\n✅ 成功生成 {row_count} 行随机数，{len(columns)} 列!")
        print(f"📁 文件已保存至: {output_path}")
        print(f"📍 数据列: {', '.join(col_letters)}，起始行: 第 {start_row} 行，结束行: 第 {start_row + row_count - 1} 行")
        print(f"🔢 随机数范围: {min_val} ~ {max_val}")
        print(f"📝 数据类型: {'整数' if is_integer else '小数(保留2位)'}")

        # 显示部分生成的数据（仅预览第一列）
        preview_col = columns[0]
        preview_letter = index_to_column_letter(preview_col)
        print(f"\n📋 生成的数据预览 (前10个，仅显示第 {preview_letter} 列):")
        preview_numbers = generate_random_numbers(min_val, max_val, min(10, row_count), is_integer)
        for i, num in enumerate(preview_numbers):
            print(f"   第 {start_row + i} 行: {num}")
        if row_count > 10:
            print(f"   ... 还有 {row_count - 10} 行数据")

        return True

    except PermissionError:
        print("❌ 文件权限错误，请确保文件未被其他程序打开")
        return False
    except Exception as e:
        print(f"❌ 处理文件时发生错误: {str(e)}")
        return False


def main():
    """
    主函数 - 交互式命令行界面
    """
    print("=" * 60)
    print("       🎲 Excel 随机数生成器 v1.1")
    print("=" * 60)
    print()
    
    # 1. 获取Excel文件路径
    print("【步骤 1/5】请输入Excel文件信息")
    print("-" * 40)
    while True:
        file_path = get_user_input("请输入Excel文件路径: ")
        if validate_file_path(file_path):
            break
    
    # 2. 获取列信息
    print("\n【步骤 2/5】请输入列信息")
    print("-" * 40)
    print("提示: 可以输入列字母(如A, B, AA)或数字(如1, 2, 27)")
    while True:
        try:
            column_input = get_user_input("请输入起始列: ")
            start_column = column_letter_to_index(column_input)
            if start_column < 1:
                print("❌ 列号必须大于0")
                continue
            col_count = get_user_input("请输入要同时生成的列数 (默认为1): ", int, 1)
            if col_count < 1:
                print("⚠️ 列数不能小于1，已自动设置为1")
                col_count = 1
            columns = [start_column + i for i in range(col_count)]
            col_letters = [index_to_column_letter(c) for c in columns]
            print(f"✓ 已选择列: {', '.join(col_letters)}")
            break
        except ValueError as e:
            print(f"❌ {str(e)}")

    # 3. 获取行信息
    print("\n【步骤 3/5】请输入行信息")
    print("-" * 40)
    mode = get_user_input("请选择起始方式: 1. 指定起始行  2. 自动接在已有数据之后 (默认为2): ", str, "2")
    auto_append = mode.strip() != "1"
    if auto_append:
        start_row = 1  # 实际起始行在处理时计算
        print("✓ 已选择自动追加模式，将接在已有数据最后一行之后")
    else:
        start_row = get_user_input("请输入起始行号 (默认为1): ", int, 1)
        if start_row < 1:
            print("⚠️ 起始行号不能小于1，已自动设置为1")
            start_row = 1
    
    row_count = get_user_input("请输入要生成的行数: ", int)
    if row_count < 1:
        print("❌ 行数必须大于0")
        return
    
    # 4. 获取随机数范围
    print("\n【步骤 4/5】请输入随机数范围")
    print("-" * 40)
    min_val = get_user_input("请输入随机数最小值 (默认为0): ", float, 0)
    max_val = get_user_input("请输入随机数最大值 (默认为100): ", float, 100)
    
    if min_val > max_val:
        print("⚠️ 最小值大于最大值，已自动交换")
        min_val, max_val = max_val, min_val
    
    # 询问是否为整数
    is_integer_input = get_user_input("是否生成整数? (y/n, 默认为n): ", str, "n")
    is_integer = is_integer_input.lower() == 'y'
    
    # 5. 确认并执行
    print("\n【步骤 5/5】确认信息")
    print("-" * 40)
    print(f"📁 文件路径: {file_path}")
    print(f"📍 目标列: {', '.join([index_to_column_letter(c) for c in columns])}")
    if auto_append:
        print("📍 起始行: 自动接在已有数据之后")
    else:
        print(f"📍 起始行: 第 {start_row} 行")
    print(f"📊 生成行数: {row_count} 行")
    print(f"🔢 随机数范围: {min_val} ~ {max_val}")
    print(f"📝 数据类型: {'整数' if is_integer else '小数(保留2位)'}")
    
    confirm = get_user_input("\n确认执行? (y/n): ", str, "y")
    if confirm.lower() != 'y':
        print("❌ 操作已取消")
        return
    
    # 询问是否保存为新文件
    save_as_new = get_user_input("是否保存为新文件? (y/n, 默认覆盖原文件): ", str, "n")
    output_path = None
    if save_as_new.lower() == 'y':
        base, ext = os.path.splitext(file_path)
        output_path = f"{base}_random{ext}"
        print(f"📁 新文件路径: {output_path}")
    
    # 执行处理
    print("\n" + "=" * 60)
    process_excel_file(file_path, columns, start_row, row_count,
                       min_val, max_val, is_integer, output_path, auto_append)
    print("=" * 60)


if __name__ == "__main__":
    main()
